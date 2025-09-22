# pip install openpyxl
import openpyxl


def read_excel_sheet(file_path, sheet_name, cell_name):
    # Load workbook from file
    wb = openpyxl.load_workbook(file_path)
    #print(wb.__dict__)
    # Get sheet name from Excel file
    sheet = wb[sheet_name]
    # Get cell name from Excel file
    cell = sheet[cell_name]
    # Get value of cell from Excel file
    print(cell.value)


read_excel_sheet(file_path="test_data.xlsx", sheet_name="Sheet1", cell_name="A1")

for i in range(1, 5):
    read_excel_sheet(file_path="test_data.xlsx", sheet_name="Sheet1", cell_name=f"A{i}")
"""
Pune
Mumbai
Bangalore
Chennai

"""

for i in range(1, 5):
    read_excel_sheet(file_path="test_data.xlsx", sheet_name="Sheet2", cell_name=f"A{i}")

"""
INDIA
CHINA
USA

"""


def write_content_to_file(file_path, sheet_name, cell_name, value):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    cell_name = sheet[cell_name]
    # update value in cell
    cell_name.value = value
    wb.save(file_path)

#write_content_to_file("test_data.xlsx", "Sheet3", "A1", "Software Engineer")

#############
# Read maximum rows and colums from excel sheet.

def read_all_rows_colums(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    max_colums = sheet.max_column
    print("max rows:", max_row)
    print("Max colums :", max_colums)
    for i in range(1, max_row+1):
        for j in range(1, max_colums+1):
            print(sheet.cell(row=i, column=j).value, end=" ")
        print()

print("_"*50)
read_all_rows_colums("test_data.xlsx", "Sheet4")

