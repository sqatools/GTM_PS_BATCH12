import openpyxl

def read_write_excelsheet(filepath, sheet_name, cell_name):
    wb= openpyxl.load_workbook(filepath)
    sheet= wb[sheet_name]
    cell= sheet[cell_name]
    print(cell.value)

for i in range(1,7):
    read_write_excelsheet("excel_practice.xlsx", "Sheet1",cell_name= f"A{i}")

print("-"*50)
def write_content_to_file(file_name, sheet_name, cell_name, value):
    wb= openpyxl.load_workbook(file_name)
    sheet= wb[sheet_name]
    cell_name= sheet[cell_name]
    cell_name.value= value
    print(cell_name.value)
    wb.save(file_name)
    wb.close()


write_content_to_file("excel_practice.xlsx", "Sheet1", cell_name="A7", value= "Python")
print("-"*50)

def read_write_excelsheet(filepath, sheet_name, cell_name):
    wb= openpyxl.load_workbook(filepath)
    sheet= wb[sheet_name]
    cell= sheet[cell_name]
    print(cell.value)

for i in range(1,8):
    read_write_excelsheet("excel_practice.xlsx", "Sheet1",cell_name= f"A{i}")




